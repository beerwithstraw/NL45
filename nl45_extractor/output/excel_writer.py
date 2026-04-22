"""
Excel Writer for NL-45 (Grievance Disposal).

Master_Data: one row per (company × quarter × complaint_type) — ten rows per
filing (nine complaint types + total). Seven status metrics as direct columns.

Verification sheet per company: mirrors PDF layout with status grid + benchmark
section. Named {CompanyPascal}_{Quarter}_{Year} (max 31 chars).

_meta sheet: extraction run metadata.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config.settings import (
    MASTER_COLUMNS, STATUS_COL_TO_METRIC, EXTRACTOR_VERSION,
    NUMBER_FORMAT,
    company_key_to_pascal, _year_code_to_fy_end,
)
from config.row_registry import NL45_ROW_ORDER, NL45_ROW_DISPLAY_NAMES
from config.company_metadata import get_metadata
from extractor.models import NL45Extract

logger = logging.getLogger(__name__)

_HEADER_FONT   = Font(bold=True, color="FFFFFF")
_HEADER_FILL   = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_SUMMARY_FONT  = Font(bold=True)
_SUMMARY_FILL  = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
_CENTER_ALIGN  = Alignment(horizontal="center", vertical="center")
_META_FILL     = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BENCH_FILL    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Column names in MASTER_COLUMNS that hold numeric metric values.
_METRIC_COLUMNS = set(STATUS_COL_TO_METRIC.keys())


# ---------------------------------------------------------------------------
# Master_Data sheet
# ---------------------------------------------------------------------------

def _write_master_data(ws, extractions: List[NL45Extract],
                       existing_rows: Optional[List[list]] = None):
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                col_name = MASTER_COLUMNS[col_idx - 1]
                if col_name in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    for extract in extractions:
        if not extract.status_data:
            continue

        meta = get_metadata(extract.company_key)
        year_display = _year_code_to_fy_end(extract.year)

        for complaint_type in NL45_ROW_ORDER:
            metrics = extract.status_data.data.get(complaint_type, {})

            is_total = (complaint_type == "total_complaints")

            row_meta = {
                "Complaint_Type":       complaint_type,
                "Complaint_Display":    NL45_ROW_DISPLAY_NAMES.get(complaint_type, complaint_type),
                "Company_Name":         meta["company_name"],
                "Company":              meta["sorted_company"],
                "NL":                   extract.form_type,
                "Quarter":              extract.quarter,
                "Year":                 year_display,
                "Sector":               meta["sector"],
                "Industry_Competitors": meta["competitors"],
                "GI_Companies":         "GI Company",
                "Source_File":          extract.source_file,
            }

            row_values = []
            for col_name in MASTER_COLUMNS:
                if col_name in row_meta:
                    row_values.append(row_meta[col_name])
                elif col_name in STATUS_COL_TO_METRIC:
                    metric_key = STATUS_COL_TO_METRIC[col_name]
                    row_values.append(metrics.get(metric_key))
                else:
                    row_values.append(None)

            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                col_name = MASTER_COLUMNS[col_idx - 1]
                if col_name in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
                if is_total:
                    cell.font = _SUMMARY_FONT
                    cell.fill = _SUMMARY_FILL

            current_row += 1


# ---------------------------------------------------------------------------
# Verification sheet — mirrors PDF layout
# ---------------------------------------------------------------------------

def _write_verification_sheet(ws, extract: NL45Extract):
    ws.cell(row=1, column=1, value=f"VERIFICATION SHEET: {extract.company_name}") \
      .font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f"Quarter: {extract.quarter} | Year: {extract.year} | Source: {extract.source_file}")

    if not extract.status_data:
        ws.cell(row=4, column=1, value="No data extracted.").font = Font(italic=True)
        return

    _write_status_section(ws, extract, start_row=4)

    if extract.benchmark_data:
        end_row = 4 + 3 + len(NL45_ROW_ORDER) + 2
        _write_benchmark_section(ws, extract.benchmark_data, start_row=end_row)


def _write_status_section(ws, extract: NL45Extract, start_row: int):
    """PDF-matching layout for Table 1."""
    col_headers = [
        "Complaint Type",
        "Opening Balance",
        "Additions",
        "Fully Accepted",
        "Partial Accepted",
        "Rejected",
        "Pending EoQ",
        "Total Registered YTD",
    ]
    metric_keys = [
        "opening_balance", "additions", "fully_accepted",
        "partial_accepted", "rejected", "pending_eoq", "total_registered_ytd",
    ]

    # Title
    title_cell = ws.cell(row=start_row, column=1, value="Grievance Disposal (Table 1 — Status)")
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = _HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row,
                   end_column=len(col_headers))
    title_cell.alignment = _CENTER_ALIGN

    # Headers
    h = start_row + 1
    for col_idx, label in enumerate(col_headers, 1):
        cell = ws.cell(row=h, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN

    ws.column_dimensions["A"].width = 30
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 20

    # Data rows
    for r_idx, complaint_type in enumerate(NL45_ROW_ORDER):
        ws_row = h + 1 + r_idx
        is_total = (complaint_type == "total_complaints")
        display = NL45_ROW_DISPLAY_NAMES.get(complaint_type, complaint_type)

        label_cell = ws.cell(row=ws_row, column=1, value=display)
        if is_total:
            label_cell.font = _SUMMARY_FONT
            label_cell.fill = _SUMMARY_FILL

        metrics = extract.status_data.data.get(complaint_type, {})
        for col_offset, metric_key in enumerate(metric_keys, 2):
            val = metrics.get(metric_key)
            cell = ws.cell(row=ws_row, column=col_offset, value=val)
            cell.number_format = NUMBER_FORMAT
            if is_total:
                cell.font = _SUMMARY_FONT
                cell.fill = _SUMMARY_FILL


def _write_benchmark_section(ws, bm, start_row: int):
    """Tabular layout for Table 2 benchmark metrics."""
    title_cell = ws.cell(row=start_row, column=1,
                         value="Benchmark / Ratio Metrics (Table 2)")
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)
    title_cell.fill = _HEADER_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    title_cell.alignment = _CENTER_ALIGN

    h = start_row + 1
    for col_idx, label in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=h, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    benchmark_rows = [
        ("Total Policies — Previous Year",           bm.policies_prev_year),
        ("Total Claims — Previous Year",             bm.claims_prev_year),
        ("Total Policies — Current Year",            bm.policies_curr_year),
        ("Total Claims — Current Year",              bm.claims_curr_year),
        ("Policy Complaints per 10,000 Policies",    bm.policy_complaints_per_10k),
        ("Claim Complaints per 10,000 Claims",       bm.claim_complaints_per_10k),
    ]

    for r_idx, (label, val) in enumerate(benchmark_rows):
        ws_row = h + 1 + r_idx
        ws.cell(row=ws_row, column=1, value=label).fill = _BENCH_FILL
        val_cell = ws.cell(row=ws_row, column=2, value=val)
        val_cell.number_format = NUMBER_FORMAT
        val_cell.fill = _BENCH_FILL


# ---------------------------------------------------------------------------
# Benchmark_Data sheet — one row per (company × quarter)
# ---------------------------------------------------------------------------

_BENCHMARK_COLUMNS = [
    "Company_Name",
    "Company",
    "NL",
    "Quarter",
    "Year",
    "Sector",
    "Total_Policies_Prev_Year",
    "Total_Claims_Prev_Year",
    "Total_Policies_Curr_Year",
    "Total_Claims_Curr_Year",
    "Policy_Complaints_Per_10k",
    "Claim_Complaints_Per_10k",
    "Source_File",
]

_BENCHMARK_FIELD_MAP = {
    "Total_Policies_Prev_Year":    "policies_prev_year",
    "Total_Claims_Prev_Year":      "claims_prev_year",
    "Total_Policies_Curr_Year":    "policies_curr_year",
    "Total_Claims_Curr_Year":      "claims_curr_year",
    "Policy_Complaints_Per_10k":   "policy_complaints_per_10k",
    "Claim_Complaints_Per_10k":    "claim_complaints_per_10k",
}


def _write_benchmark_data(ws, extractions: List[NL45Extract],
                          existing_rows: Optional[List[list]] = None):
    for col_idx, col_name in enumerate(_BENCHMARK_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN
    ws.freeze_panes = "A2"

    current_row = 2

    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(_BENCHMARK_COLUMNS):
                    break
                ws.cell(row=current_row, column=col_idx, value=val)
            current_row += 1

    for extract in extractions:
        if not extract.benchmark_data:
            continue

        meta = get_metadata(extract.company_key)
        bm = extract.benchmark_data
        year_display = _year_code_to_fy_end(extract.year)

        row_meta = {
            "Company_Name": meta["company_name"],
            "Company":      meta["sorted_company"],
            "NL":           extract.form_type,
            "Quarter":      extract.quarter,
            "Year":         year_display,
            "Sector":       meta["sector"],
            "Source_File":  extract.source_file,
        }

        row_values = []
        for col_name in _BENCHMARK_COLUMNS:
            if col_name in row_meta:
                row_values.append(row_meta[col_name])
            elif col_name in _BENCHMARK_FIELD_MAP:
                field = _BENCHMARK_FIELD_MAP[col_name]
                row_values.append(getattr(bm, field, None))
            else:
                row_values.append(None)

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            if _BENCHMARK_COLUMNS[col_idx - 1] in _BENCHMARK_FIELD_MAP:
                cell.number_format = NUMBER_FORMAT

        current_row += 1


# ---------------------------------------------------------------------------
# Meta sheet
# ---------------------------------------------------------------------------

def _write_meta_sheet(ws, extractions: List[NL45Extract], stats: Dict[str, Any]):
    companies = sorted({e.company_name for e in extractions})
    quarters  = sorted({f"{e.quarter}_{e.year}" for e in extractions})

    data = [
        ["Key", "Value"],
        ["extraction_date",   datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version", EXTRACTOR_VERSION],
        ["files_processed",   stats.get("files_processed", 0)],
        ["files_succeeded",   stats.get("files_succeeded", 0)],
        ["files_failed",      stats.get("files_failed", 0)],
        ["companies",         ", ".join(companies)],
        ["quarters",          ", ".join(quarters)],
    ]
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL


def _sheet_name_for(extract: NL45Extract) -> str:
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def save_workbook(extractions: List[NL45Extract], output_path: str,
                  stats: Optional[Dict[str, Any]] = None,
                  force: bool = False):
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []
    existing_bm_rows = []

    if output_file.exists() and not force:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)
        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]
            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                # Deduplicate by (Company, Complaint_Type, Quarter, Year) — temp filenames change each run
                try:
                    co_idx = headers.index("Company")
                    ct_idx = headers.index("Complaint_Type")
                    qt_idx = headers.index("Quarter")
                    yr_idx = headers.index("Year")
                except ValueError:
                    co_idx = ct_idx = qt_idx = yr_idx = None
                new_md_keys = {
                    (get_metadata(e.company_key)["sorted_company"], ct, e.quarter,
                     _year_code_to_fy_end(e.year))
                    for e in extractions if e.status_data
                    for ct in e.status_data.data.keys()
                }
                if co_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        key = (row[co_idx], row[ct_idx], row[qt_idx], row[yr_idx])
                        if key not in new_md_keys:
                            existing_rows.append(list(row))
            else:
                logger.warning("Existing Master_Data has different column layout — regenerating.")
            del wb["Master_Data"]

        if "Benchmark_Data" in wb.sheetnames:
            ws_bm_old = wb["Benchmark_Data"]
            bm_headers = [cell.value for cell in ws_bm_old[1]]
            if bm_headers[:len(_BENCHMARK_COLUMNS)] == _BENCHMARK_COLUMNS:
                # Deduplicate by (Company, Quarter, Year) — temp filenames change every run
                try:
                    co_idx = bm_headers.index("Company")
                    qt_idx = bm_headers.index("Quarter")
                    yr_idx = bm_headers.index("Year")
                except ValueError:
                    co_idx = qt_idx = yr_idx = None
                new_bm_keys = {
                    (get_metadata(e.company_key)["sorted_company"], e.quarter,
                     _year_code_to_fy_end(e.year))
                    for e in extractions if e.benchmark_data
                }
                if co_idx is not None:
                    for row in ws_bm_old.iter_rows(min_row=2, values_only=True):
                        key = (row[co_idx], row[qt_idx], row[yr_idx])
                        if key not in new_bm_keys:
                            existing_bm_rows.append(list(row))
            del wb["Benchmark_Data"]

        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]
        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        # force=True — always start with a fresh workbook
        wb = Workbook()
        wb.remove(wb.active)

    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows)

    ws_bm = wb.create_sheet("Benchmark_Data", 1)
    _write_benchmark_data(ws_bm, extractions, existing_rows=existing_bm_rows)

    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")


def write_validation_summary_sheet(report_path: str, master_path: str, force_company: str = None):
    import pandas as pd
    df = pd.read_csv(report_path)
    summary = df.pivot_table(
        index=["company", "quarter", "year"],
        columns="status", aggfunc="size", fill_value=0,
    ).reset_index()
    for col in ["PASS", "WARN", "FAIL", "SKIP"]:
        if col not in summary.columns:
            summary[col] = 0
    summary["Total_Checks"] = summary[["PASS", "SKIP", "WARN", "FAIL"]].sum(axis=1)
    summary = summary.rename(columns={"company": "Company", "quarter": "Quarter", "year": "Year"})
    cols = ["Company", "Quarter", "Year", "Total_Checks", "PASS", "SKIP", "WARN", "FAIL"]
    summary = summary[cols]
    if force_company:
        try:
            existing = pd.read_excel(master_path, sheet_name="Validation_Summary")
            companies_in_new = set(summary["Company"].unique())
            existing = existing[~existing["Company"].isin(companies_in_new)]
            summary = pd.concat([existing, summary], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        summary.to_excel(w, sheet_name="Validation_Summary", index=False)


def write_validation_detail_sheet(report_path: str, master_path: str, force_company: str = None):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)
    cols_map = {
        "company": "Company", "quarter": "Quarter", "year": "Year",
        "complaint_type": "Complaint_Type", "check_name": "Check_Name",
        "status": "Status", "expected": "Expected", "actual": "Actual",
        "delta": "Delta", "note": "Note",
    }
    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        rename_cols = {k: v for k, v in cols_map.items() if k in detail.columns}
        detail = detail.rename(columns=rename_cols)
        final_cols = [v for v in cols_map.values() if v in detail.columns]
        detail = detail[final_cols].sort_values("Status").reset_index(drop=True)

    if force_company:
        try:
            run_companies = set(pd.read_csv(report_path)["company"].unique())
            existing_detail = pd.read_excel(master_path, sheet_name="Validation_Detail")
            if "Company" in existing_detail.columns:
                existing_detail = existing_detail[~existing_detail["Company"].isin(run_companies)]
            detail = pd.concat([existing_detail, detail], ignore_index=True)
        except Exception:
            pass
    with pd.ExcelWriter(master_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as w:
        detail.to_excel(w, sheet_name="Validation_Detail", index=False)

    wb = load_workbook(master_path)
    ws = wb["Validation_Detail"]
    red_fill    = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    status_col_vals = list(cols_map.values())
    if "Status" in status_col_vals:
        status_col = status_col_vals.index("Status") + 1
        for row_idx in range(2, ws.max_row + 1):
            val  = ws.cell(row=row_idx, column=status_col).value
            fill = red_fill if val == "FAIL" else yellow_fill
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")
